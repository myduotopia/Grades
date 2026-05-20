"""Grade batch-import via Excel.

Two endpoints:
- GET  /api/classrooms/{id}/grades/template.xlsx — blank template (one column)
- POST /api/classrooms/{id}/grades/import          — two-phase preview + commit

The Excel file does NOT carry subject info; the teacher picks it per column
in the preview UI and sends the choices back on commit via the `subjects`
multipart form field (JSON string).
"""
from __future__ import annotations

import json
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Annotated, Any
from uuid import UUID

from fastapi import (
    APIRouter,
    Depends,
    File,
    Form,
    HTTPException,
    Query,
    UploadFile,
    status,
)
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy.orm import Session

from auth import require_user_id
from database import get_db
from models.classroom import Classroom, Student
from models.curriculum import (
    Category,
    ClassroomItem,
    Item,
    Semester,
    Subject,
    SubjectCategoryWeight,
)
from sqlalchemy.dialects.postgresql import insert as pg_insert
from models.grading import Grade, PointRecord, StudentStandard, SubjectPointRule
from schemas import (
    ClassroomGradesView,
    GradeEntryOut,
    GradeImportColumnPreview,
    GradeImportPreviewSummary,
    GradeImportResult,
    GradeImportStudentRow,
    ItemOut,
    SemesterOut,
    StudentBriefOut,
    SubjectCategoryWeightOut,
)

router = APIRouter()


# ---------- constants ----------

# Grade import is for the 3 exam-shaped categories only. 出席率 (attendance)
# and 額外加分 (extra) have dedicated UIs because they don't fit the
# "one exam = one score per student" shape.
# Categories that trigger auto-award when score >= the student's standard for
# the item's subject. Other categories (homework / attendance / extra) never
# award points automatically.
AUTO_AWARD_CATEGORY_KEYS: frozenset[str] = frozenset({"major_exam", "quiz"})


_CATEGORY_NAME_TO_KEY: dict[str, str] = {
    "段考": "major_exam",
    "小考": "quiz",
    "作業": "homework",
    "Major Exam": "major_exam",
    "Quiz": "quiz",
    "Homework": "homework",
}

# Inverse: system category key → Chinese label. Used by apply_auto_award to
# format the PointRecord.reason as "達成標準 - 段考: 第一次段考".
_CATEGORY_KEY_TO_NAME_ZH: dict[str, str] = {
    "major_exam": "段考",
    "quiz": "小考",
    "homework": "作業",
    "attendance": "出席",
    "extra": "加分",
}

MEETING_STANDARD_REASON = "達成標準"

# Excel layout: 3 metadata rows above the scores.
#   Row 1: 類別 (dropdown)
#   Row 2: 日期 (選填)
#   Row 3: 考試名稱 (選填)
#   Row 4+: scores
_DATA_ROW = 4


# ---------- error helpers ----------

def _not_found(resource: str) -> HTTPException:
    return HTTPException(
        status_code=status.HTTP_404_NOT_FOUND,
        detail={
            "error": {
                "code": "NOT_FOUND",
                "message_key": f"errors.{resource}.not_found",
                "message": f"{resource.capitalize()} not found.",
            }
        },
    )


def _bad_request(message_key: str, message: str) -> HTTPException:
    return HTTPException(
        status_code=status.HTTP_400_BAD_REQUEST,
        detail={
            "error": {
                "code": "BAD_REQUEST",
                "message_key": message_key,
                "message": message,
            }
        },
    )


# ---------- lookups ----------

def _get_owned_classroom(db: Session, user_id: UUID, classroom_id: UUID) -> Classroom:
    c = (
        db.query(Classroom)
        .filter(Classroom.id == classroom_id, Classroom.user_id == user_id)
        .one_or_none()
    )
    if c is None:
        raise _not_found("classroom")
    return c


# ---------- template ----------

@router.get(
    "/api/classrooms/{classroom_id}/grades/template.xlsx",
    response_class=StreamingResponse,
)
def download_template(
    classroom_id: UUID,
    user_id: Annotated[UUID, Depends(require_user_id)],
    db: Annotated[Session, Depends(get_db)],
) -> StreamingResponse:
    _get_owned_classroom(db, user_id, classroom_id)
    wb = Workbook()
    ws = wb.active
    ws.title = "grades"

    # Header column (A) — seat number marker
    ws.cell(row=1, column=1, value="座號")
    # Example score column (B)
    ws.cell(row=1, column=2, value="段考")  # category (dropdown)
    ws.cell(row=2, column=2, value="2026-05-13")  # date (optional)
    ws.cell(row=3, column=2, value="期中考")  # exam name (optional)
    # Example student rows (teacher deletes before uploading their own)
    for i in range(4):
        ws.cell(row=_DATA_ROW + i, column=1, value=i + 1)
    ws.cell(row=_DATA_ROW, column=2, value=80)
    ws.cell(row=_DATA_ROW + 1, column=2, value=88)
    ws.cell(row=_DATA_ROW + 2, column=2, value=60)
    ws.cell(row=_DATA_ROW + 3, column=2, value=95)

    # Category dropdown for B1:Z1 so adding more columns auto-applies.
    dv = DataValidation(
        type="list",
        formula1='"段考,小考,作業"',
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="無效的類別",
        error="只能選 段考 / 小考 / 作業",
    )
    dv.add("B1:Z1")
    ws.add_data_validation(dv)

    ws.column_dimensions["A"].width = 10
    for col_letter in ("B", "C", "D", "E", "F", "G"):
        ws.column_dimensions[col_letter].width = 16

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": 'attachment; filename="grades_template.xlsx"',
        },
    )


# ---------- import (preview + commit) ----------

@router.post(
    "/api/classrooms/{classroom_id}/grades/import",
    response_model=GradeImportResult,
)
async def import_grades(
    classroom_id: UUID,
    user_id: Annotated[UUID, Depends(require_user_id)],
    db: Annotated[Session, Depends(get_db)],
    file: Annotated[UploadFile, File(...)],
    dry_run: Annotated[bool, Query()] = True,
    # JSON string: { "<column_index>": "<subject_system_key>" }. Required for
    # commit (dry_run=false); ignored for preview.
    subjects: Annotated[str | None, Form()] = None,
) -> GradeImportResult:
    classroom = _get_owned_classroom(db, user_id, classroom_id)

    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise _bad_request("errors.import.bad_file_type", "File must be .xlsx")

    content = await file.read()
    try:
        wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise _bad_request(
            "errors.import.unreadable", "Excel file could not be parsed."
        )
    ws = wb.active
    if ws is None:
        raise _bad_request("errors.import.empty", "Workbook is empty.")

    if str(ws.cell(row=1, column=1).value or "").strip() != "座號":
        raise _bad_request(
            "errors.import.bad_header", "A1 must be 「座號」"
        )

    current_semester = (
        db.query(Semester)
        .filter(Semester.user_id == user_id, Semester.is_current.is_(True))
        .one_or_none()
    )
    if current_semester is None:
        raise _bad_request(
            "errors.import.no_current_semester",
            "No current semester is set. Configure it before importing grades.",
        )

    roster = (
        db.query(Student)
        .filter(Student.classroom_id == classroom_id, Student.user_id == user_id)
        .all()
    )
    by_seat: dict[int, Student] = {s.seat_number: s for s in roster}

    columns = _parse_columns(ws)
    student_rows = _parse_rows(ws, by_seat, columns)

    score_count = sum(len(r.scores) for r in student_rows if not r.errors)
    error_count = sum(1 for c in columns if c.errors) + sum(
        1 for r in student_rows if r.errors
    )

    summary = GradeImportPreviewSummary(
        column_total=len(columns),
        row_total=len(student_rows),
        score_total=score_count,
        errors=error_count,
    )

    if dry_run:
        return GradeImportResult(
            dry_run=True, summary=summary, columns=columns, students=student_rows,
        )

    # Commit phase — subjects required and no errors allowed.
    if error_count > 0:
        raise _bad_request(
            "errors.import.has_errors",
            "Cannot import while preview contains errors.",
        )
    if subjects is None:
        raise _bad_request(
            "errors.import.subject_required_for_commit",
            "Subject choices are required to commit.",
        )
    try:
        subject_map_raw: dict[str, str] = json.loads(subjects)
    except (ValueError, TypeError):
        raise _bad_request(
            "errors.import.bad_subjects_payload",
            "`subjects` form field must be valid JSON.",
        )
    # Normalize keys to int (frontend sends column_index as string).
    subject_map: dict[int, str] = {}
    for k, v in subject_map_raw.items():
        try:
            subject_map[int(k)] = str(v)
        except (TypeError, ValueError):
            raise _bad_request(
                "errors.import.bad_subjects_payload",
                "`subjects` keys must be column indexes.",
            )

    _commit_grades(
        db, user_id, classroom, current_semester, columns, student_rows, subject_map,
    )
    db.commit()
    return GradeImportResult(
        dry_run=False, summary=summary, columns=columns, students=student_rows,
    )


# ---------- parse ----------

def _parse_columns(ws: Any) -> list[GradeImportColumnPreview]:
    """Walk columns B..end. Each non-blank column = one Item to write."""
    result: list[GradeImportColumnPreview] = []
    max_col = ws.max_column or 0
    for col in range(2, max_col + 1):
        category_raw = ws.cell(row=1, column=col).value
        date_raw = ws.cell(row=2, column=col).value
        name_raw = ws.cell(row=3, column=col).value

        # Skip entirely-blank columns (no category AND no scores).
        if category_raw in (None, ""):
            has_scores = any(
                ws.cell(row=r, column=col).value not in (None, "")
                for r in range(_DATA_ROW, (ws.max_row or _DATA_ROW) + 1)
            )
            if not has_scores:
                continue

        errors: list[str] = []
        category_str = (
            str(category_raw).strip() if category_raw not in (None, "") else None
        )
        category_key = _CATEGORY_NAME_TO_KEY.get(category_str) if category_str else None
        if category_str is None:
            errors.append("類別 is required")
        elif category_key is None:
            errors.append(f"Unknown 類別: {category_str} (allowed: 段考 / 小考 / 作業)")

        exam_date = _coerce_date(date_raw)
        if date_raw not in (None, "") and exam_date is None:
            errors.append(f"日期格式無法解析: {date_raw!r}")
        if exam_date is None:
            exam_date = date.today()

        if name_raw not in (None, ""):
            exam_name = str(name_raw).strip()
        else:
            cat_label = category_str or ""
            exam_name = f"{cat_label}-{exam_date.isoformat()}"

        result.append(
            GradeImportColumnPreview(
                column_index=col - 1,  # 0-based; B = 1
                category_input=category_str,
                category_system_key=category_key,
                exam_date=exam_date,
                exam_name=exam_name,
                errors=errors,
            )
        )
    return result


def _parse_rows(
    ws: Any,
    by_seat: dict[int, Student],
    columns: list[GradeImportColumnPreview],
) -> list[GradeImportStudentRow]:
    rows: list[GradeImportStudentRow] = []
    seen_seats: set[int] = set()
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    for r in range(_DATA_ROW, max_row + 1):
        cells = [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
        if all(v in (None, "") for v in cells):
            continue

        errors: list[str] = []
        seat_raw = ws.cell(row=r, column=1).value
        seat: int | None = None
        try:
            if seat_raw in (None, ""):
                errors.append("座號 is required")
            else:
                seat = int(seat_raw)
                if seat < 1 or seat > 99:
                    errors.append("座號 must be 1–99")
                    seat = None
        except (TypeError, ValueError):
            errors.append("座號 must be an integer")

        if seat is not None:
            if seat in seen_seats:
                errors.append("座號 duplicated in file")
            seen_seats.add(seat)

        student = by_seat.get(seat) if seat is not None else None
        if seat is not None and student is None:
            errors.append(
                f"座號 {seat} not found in this classroom — import roster first"
            )

        scores: dict[int, float] = {}
        for col in columns:
            if col.errors:
                continue
            cell_val = ws.cell(row=r, column=col.column_index + 1).value
            if cell_val in (None, ""):
                continue
            try:
                num = float(Decimal(str(cell_val)))
            except (InvalidOperation, ValueError):
                errors.append(
                    f"col {col.column_index + 1} is not a number"
                )
                continue
            if num < 0 or num > 100:
                errors.append(f"col {col.column_index + 1} must be 0–100")
                continue
            scores[col.column_index] = num

        rows.append(
            GradeImportStudentRow(
                row_number=r,
                seat_number=seat,
                student_id=student.id if student else None,
                scores=scores,
                errors=errors,
            )
        )
    return rows


def _coerce_date(raw: Any) -> date | None:
    if raw is None or raw == "":
        return None
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    if isinstance(raw, int) and 19000101 <= raw <= 99991231:
        s = str(raw)
        try:
            return date(int(s[0:4]), int(s[4:6]), int(s[6:8]))
        except ValueError:
            return None
    if isinstance(raw, str):
        s = raw.strip().replace("/", "-")
        for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%Y%m%d"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
    return None


# ---------- commit ----------

def _commit_grades(
    db: Session,
    user_id: UUID,
    classroom: Classroom,
    semester: Semester,
    columns: list[GradeImportColumnPreview],
    student_rows: list[GradeImportStudentRow],
    subject_map: dict[int, str],
) -> None:
    # Reference data
    subjects_by_key: dict[str, Subject] = {
        s.system_key: s
        for s in db.query(Subject)
        .filter(Subject.user_id.is_(None), Subject.system_key.isnot(None))
        .all()
    }
    categories_by_key: dict[str, Category] = {
        c.system_key: c
        for c in db.query(Category).filter(Category.user_id == user_id).all()
    }

    # Validate the subject map covers every non-error column.
    missing = [c.column_index for c in columns if not c.errors and c.column_index not in subject_map]
    if missing:
        raise _bad_request(
            "errors.import.subject_required_for_commit",
            f"Pick a subject for every column (missing: {missing})",
        )

    # Upsert one Item per column.
    col_to_item: dict[int, Item] = {}
    for col in columns:
        if col.errors:
            continue
        subject_key = subject_map[col.column_index]
        subj = subjects_by_key.get(subject_key)
        if subj is None:
            raise _bad_request(
                "errors.import.subject_invalid",
                f"Unknown subject system_key: {subject_key}",
            )
        cat = categories_by_key.get(col.category_system_key or "")
        if cat is None:
            raise _bad_request(
                "errors.import.category_not_seeded",
                f"Category {col.category_system_key} not seeded for user.",
            )
        existing = (
            db.query(Item)
            .filter(
                Item.user_id == user_id,
                Item.subject_id == subj.id,
                Item.category_id == cat.id,
                Item.semester_id == semester.id,
                Item.name == col.exam_name,
            )
            .one_or_none()
        )
        if existing is not None:
            item = existing
        else:
            item = Item(
                user_id=user_id,
                subject_id=subj.id,
                category_id=cat.id,
                semester_id=semester.id,
                name=col.exam_name,
            )
            db.add(item)
            db.flush()
        col_to_item[col.column_index] = item
    db.flush()

    # Activate every imported item for this classroom (idempotent).
    for item in col_to_item.values():
        db.execute(
            pg_insert(ClassroomItem)
            .values(
                user_id=user_id,
                classroom_id=classroom_id,
                item_id=item.id,
            )
            .on_conflict_do_nothing(constraint="uq_classroom_item")
        )

    # Upsert grades.
    written_grades: list[Grade] = []
    for r in student_rows:
        if r.errors or r.student_id is None:
            continue
        for col_idx, score in r.scores.items():
            item = col_to_item.get(col_idx)
            if item is None:
                continue
            existing_g = (
                db.query(Grade)
                .filter(Grade.item_id == item.id, Grade.student_id == r.student_id)
                .one_or_none()
            )
            if existing_g is not None:
                existing_g.score = Decimal(str(score))
                written_grades.append(existing_g)
            else:
                g = Grade(
                    user_id=user_id,
                    item_id=item.id,
                    student_id=r.student_id,
                    score=Decimal(str(score)),
                    source="manual",
                )
                db.add(g)
                written_grades.append(g)
    db.flush()
    apply_auto_award(db, user_id, written_grades)


def apply_auto_award(
    db: Session, user_id: UUID, grades: list[Grade]
) -> None:
    """Award / revoke point_records for the given grades in-place.

    For each grade:
      1. If item.category.system_key NOT IN AUTO_AWARD_CATEGORY_KEYS → no-op.
      2. Look up StudentStandard(student_id, item.subject_id).threshold;
         if absent → no-op.
      3. Look up SubjectPointRule(user_id, item.subject_id).points_awarded.
      4. If score >= threshold: upsert PointRecord(source_grade_id=grade.id).
      5. If score < threshold: delete any existing PointRecord with that
         source_grade_id (revoke an earlier award).
    """
    if not grades:
        return

    item_ids = {g.item_id for g in grades}
    items_by_id = {
        i.id: i
        for i in db.query(Item).filter(Item.id.in_(item_ids)).all()
    }
    cat_ids = {it.category_id for it in items_by_id.values()}
    cat_keys: dict[UUID, str] = {
        c.id: c.system_key
        for c in db.query(Category).filter(Category.id.in_(cat_ids)).all()
    }
    subj_ids = {it.subject_id for it in items_by_id.values()}
    subj_labels: dict[UUID, str] = {
        s.id: (s.system_key or s.display_name or "")
        for s in db.query(Subject).filter(Subject.id.in_(subj_ids)).all()
    }
    point_rule_rows = (
        db.query(SubjectPointRule)
        .filter(
            SubjectPointRule.user_id == user_id,
            SubjectPointRule.subject_id.in_(subj_ids),
        )
        .all()
    )
    points_by_subject: dict[UUID, int] = {
        r.subject_id: r.points_awarded for r in point_rule_rows
    }

    student_ids = {g.student_id for g in grades}
    standards = (
        db.query(StudentStandard)
        .filter(
            StudentStandard.user_id == user_id,
            StudentStandard.student_id.in_(student_ids),
            StudentStandard.subject_id.in_(subj_ids),
        )
        .all()
    )
    standard_by_pair: dict[tuple[UUID, UUID], Decimal] = {
        (s.student_id, s.subject_id): s.threshold for s in standards
    }

    grade_ids = [g.id for g in grades]
    existing_records = (
        db.query(PointRecord)
        .filter(PointRecord.source_grade_id.in_(grade_ids))
        .all()
    )
    record_by_grade: dict[UUID, PointRecord] = {
        r.source_grade_id: r for r in existing_records if r.source_grade_id
    }

    for g in grades:
        item = items_by_id.get(g.item_id)
        if item is None:
            continue
        cat_key = cat_keys.get(item.category_id, "")
        if cat_key not in AUTO_AWARD_CATEGORY_KEYS:
            continue
        threshold = standard_by_pair.get((g.student_id, item.subject_id))
        if threshold is None:
            continue
        pts = points_by_subject.get(item.subject_id)
        if pts is None or pts <= 0:
            # No rule or zero points configured — revoke any earlier award.
            rec = record_by_grade.get(g.id)
            if rec is not None:
                db.delete(rec)
            continue

        meets = g.score >= threshold
        rec = record_by_grade.get(g.id)
        if meets:
            cat_label = _CATEGORY_KEY_TO_NAME_ZH.get(cat_key, cat_key)
            item_label = (item.name or "").strip() or cat_label
            reason = f"{MEETING_STANDARD_REASON} - {cat_label}: {item_label}"
            if rec is None:
                db.add(
                    PointRecord(
                        user_id=user_id,
                        student_id=g.student_id,
                        points=pts,
                        reason=reason,
                        source_grade_id=g.id,
                    )
                )
            else:
                rec.points = pts
                rec.reason = reason
        else:
            if rec is not None:
                db.delete(rec)


# ---------- deactivate an item for a classroom ----------

@router.delete(
    "/api/classrooms/{classroom_id}/items/{item_id}/activation",
    status_code=status.HTTP_204_NO_CONTENT,
)
def deactivate_classroom_item(
    classroom_id: UUID,
    item_id: UUID,
    user_id: Annotated[UUID, Depends(require_user_id)],
    db: Annotated[Session, Depends(get_db)],
) -> None:
    """Hide an item from the classroom's grade view. Historical Grade rows
    (and the underlying Item) are preserved — re-activating later via the
    online grade-entry save flow brings the column back with old grades
    intact.

    Guard: rejects 409 if any student in this classroom still has a
    score > 0 for this item. Forces the teacher to explicitly clear /
    zero-out scores via the entry page before they can hide the column —
    prevents accidentally orphaning a class's worth of real grades.
    Grades with score = 0 don't block (they're "no real result")."""
    _get_owned_classroom(db, user_id, classroom_id)

    has_real_score = (
        db.query(Grade.id)
        .join(Student, Student.id == Grade.student_id)
        .filter(
            Student.classroom_id == classroom_id,
            Grade.item_id == item_id,
            Grade.score > 0,
        )
        .first()
        is not None
    )
    if has_real_score:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail={
                "error": {
                    "code": "CONFLICT",
                    "message_key": "errors.classroom_item.has_scores",
                    "message": "Cannot deactivate: some students still have scores > 0. Clear them first.",
                }
            },
        )

    db.query(ClassroomItem).filter(
        ClassroomItem.user_id == user_id,
        ClassroomItem.classroom_id == classroom_id,
        ClassroomItem.item_id == item_id,
    ).delete()
    db.commit()
    return None


# ---------- view: classroom grades bundle ----------

@router.get(
    "/api/classrooms/{classroom_id}/grades",
    response_model=ClassroomGradesView,
)
def get_classroom_grades(
    classroom_id: UUID,
    user_id: Annotated[UUID, Depends(require_user_id)],
    db: Annotated[Session, Depends(get_db)],
    semester_id: UUID | None = None,
) -> ClassroomGradesView:
    """Return one classroom's full grade bundle for one semester.

    Frontend computes weighted totals from `category_weights` × `grades`. The
    payload is bounded by classroom size × items per semester — small enough
    to ship as a single response.
    """
    classroom = _get_owned_classroom(db, user_id, classroom_id)

    if semester_id is not None:
        semester = (
            db.query(Semester)
            .filter(Semester.id == semester_id, Semester.user_id == user_id)
            .one_or_none()
        )
        if semester is None:
            raise _not_found("semester")
    else:
        semester = (
            db.query(Semester)
            .filter(Semester.user_id == user_id, Semester.is_current.is_(True))
            .one_or_none()
        )
        if semester is None:
            raise _bad_request(
                "errors.import.no_current_semester",
                "No current semester is set.",
            )

    # Category id ↔ system_key lookup for resolving item categories below.
    cats = db.query(Category).filter(Category.user_id == user_id).all()
    cat_id_to_key: dict[UUID, str] = {c.id: c.system_key for c in cats}

    # All subjects visible to the user (built-in + custom).
    subjects = (
        db.query(Subject)
        .filter((Subject.user_id.is_(None)) | (Subject.user_id == user_id))
        .all()
    )
    subj_by_id = {s.id: s for s in subjects}

    # Per-subject category weights for this user.
    scw_rows = (
        db.query(SubjectCategoryWeight)
        .filter(SubjectCategoryWeight.user_id == user_id)
        .all()
    )
    subject_category_weights = [
        SubjectCategoryWeightOut(
            subject_id=r.subject_id,
            subject_system_key=(
                subj_by_id[r.subject_id].system_key
                if r.subject_id in subj_by_id
                else None
            ),
            subject_display_name=(
                subj_by_id[r.subject_id].display_name
                if r.subject_id in subj_by_id
                else None
            ),
            category_system_key=cat_id_to_key.get(r.category_id, ""),
            weight=r.weight,
        )
        for r in scw_rows
    ]

    # Roster
    students = (
        db.query(Student)
        .filter(
            Student.classroom_id == classroom_id, Student.user_id == user_id
        )
        .order_by(Student.seat_number.asc())
        .all()
    )
    student_ids = {s.id for s in students}

    # Items are cross-classroom (one Item row per quiz across all classes),
    # but each classroom's view only includes items it has "activated" —
    # via the online grade-entry save flow or by importing grades. See the
    # ClassroomItem model. Teachers can deactivate a column from the column
    # header; the underlying Item and historical Grade rows are preserved.
    items = (
        db.query(Item)
        .join(
            ClassroomItem,
            (ClassroomItem.item_id == Item.id)
            & (ClassroomItem.classroom_id == classroom_id),
        )
        .filter(
            Item.user_id == user_id,
            Item.semester_id == semester.id,
        )
        .all()
    )

    item_outs = [
        ItemOut(
            id=i.id,
            name=i.name,
            subject_id=i.subject_id,
            subject_system_key=(
                subj_by_id[i.subject_id].system_key
                if i.subject_id in subj_by_id
                else None
            ),
            subject_display_name=(
                subj_by_id[i.subject_id].display_name
                if i.subject_id in subj_by_id
                else None
            ),
            category_system_key=cat_id_to_key.get(i.category_id, ""),
        )
        for i in items
    ]

    # Grades for those items × this roster
    item_ids = [i.id for i in items]
    grades: list[Grade] = []
    if item_ids and student_ids:
        grades = (
            db.query(Grade)
            .filter(
                Grade.item_id.in_(item_ids),
                Grade.student_id.in_(student_ids),
            )
            .all()
        )

    return ClassroomGradesView(
        semester=SemesterOut.model_validate(semester),
        subject_category_weights=subject_category_weights,
        students=[StudentBriefOut.model_validate(s) for s in students],
        items=item_outs,
        grades=[
            GradeEntryOut(
                item_id=g.item_id,
                student_id=g.student_id,
                score=float(g.score),
            )
            for g in grades
        ],
    )

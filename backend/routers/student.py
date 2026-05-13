"""Student CRUD, Excel batch import, and template download.

All routes are scoped to the authenticated user. Classroom-nested routes
(list/create/import/template) live under /api/classrooms/{id}/students; the
entity-level routes (update/delete) live under /api/students/{id}.
"""
from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Annotated, Any, Literal
from uuid import UUID

from fastapi import (
    APIRouter,
    Depends,
    File,
    HTTPException,
    Query,
    UploadFile,
    status,
)
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session, selectinload

from auth import require_user_id
from database import get_db
from models.classroom import Classroom, Student
from models.curriculum import Category, Item, Semester, Subject
from models.grading import Grade, StudentStandard
from schemas import (
    CATEGORY_NAME_TO_KEY,
    SUBJECT_NAME_TO_KEY,
    ImportColumnPreview,
    ImportPreviewSummary,
    ImportResult,
    ImportStudentRow,
    ListMeta,
    StudentCreate,
    StudentList,
    StudentOut,
    StudentStandardOut,
    StudentUpdate,
)

router = APIRouter()


# ---------- error helpers ----------

def _not_found(resource: str) -> HTTPException:
    return HTTPException(
        status_code=status.HTTP_404_NOT_FOUND,
        detail={
            "error": {
                "code": "NOT_FOUND",
                "message_key": f"errors.{resource}.not_found",
                "message": f"{resource.capitalize()} not found.",
            }
        },
    )


def _conflict(message_key: str, message: str) -> HTTPException:
    return HTTPException(
        status_code=status.HTTP_409_CONFLICT,
        detail={
            "error": {
                "code": "CONFLICT",
                "message_key": message_key,
                "message": message,
            }
        },
    )


def _bad_request(message_key: str, message: str) -> HTTPException:
    return HTTPException(
        status_code=status.HTTP_400_BAD_REQUEST,
        detail={
            "error": {
                "code": "BAD_REQUEST",
                "message_key": message_key,
                "message": message,
            }
        },
    )


# ---------- lookups ----------

def _get_owned_classroom(db: Session, user_id: UUID, classroom_id: UUID) -> Classroom:
    c = (
        db.query(Classroom)
        .filter(Classroom.id == classroom_id, Classroom.user_id == user_id)
        .one_or_none()
    )
    if c is None:
        raise _not_found("classroom")
    return c


def _get_owned_student(db: Session, user_id: UUID, student_id: UUID) -> Student:
    s = (
        db.query(Student)
        .filter(Student.id == student_id, Student.user_id == user_id)
        .one_or_none()
    )
    if s is None:
        raise _not_found("student")
    return s


def _categories_by_key(db: Session, user_id: UUID) -> dict[str, Category]:
    rows = db.query(Category).filter(Category.user_id == user_id).all()
    return {c.system_key: c for c in rows}


def _serialize_student(s: Student, cats_by_id: dict[UUID, str]) -> StudentOut:
    return StudentOut(
        id=s.id,
        classroom_id=s.classroom_id,
        seat_number=s.seat_number,
        name=s.name,
        email=s.email,
        source=s.source,
        created_at=s.created_at,
        updated_at=s.updated_at,
        standards=[
            StudentStandardOut(
                system_key=cats_by_id.get(std.category_id, ""),
                threshold=float(std.threshold),
            )
            for std in s.standards
            if std.category_id in cats_by_id
        ],
    )


# ---------- CRUD ----------

@router.get("/api/classrooms/{classroom_id}/students", response_model=StudentList)
def list_students(
    classroom_id: UUID,
    user_id: Annotated[UUID, Depends(require_user_id)],
    db: Annotated[Session, Depends(get_db)],
) -> StudentList:
    _get_owned_classroom(db, user_id, classroom_id)
    rows = (
        db.query(Student)
        .options(selectinload(Student.standards))
        .filter(Student.classroom_id == classroom_id, Student.user_id == user_id)
        .order_by(Student.seat_number.asc())
        .all()
    )
    cats = _categories_by_key(db, user_id)
    cats_by_id = {c.id: k for k, c in cats.items()}
    return StudentList(
        data=[_serialize_student(s, cats_by_id) for s in rows],
        meta=ListMeta(total=len(rows)),
    )


@router.post(
    "/api/classrooms/{classroom_id}/students",
    response_model=StudentOut,
    status_code=status.HTTP_201_CREATED,
)
def create_student(
    classroom_id: UUID,
    body: StudentCreate,
    user_id: Annotated[UUID, Depends(require_user_id)],
    db: Annotated[Session, Depends(get_db)],
) -> StudentOut:
    _get_owned_classroom(db, user_id, classroom_id)
    student = Student(
        user_id=user_id,
        classroom_id=classroom_id,
        seat_number=body.seat_number,
        name=body.name,
        email=body.email,
        source="manual",
    )
    db.add(student)
    try:
        db.flush()
    except IntegrityError:
        db.rollback()
        raise _conflict(
            "errors.student.duplicate_seat",
            "Seat number already exists in this classroom.",
        )
    _apply_standards(db, user_id, student, body.standards or {})
    db.commit()
    db.refresh(student)
    cats = _categories_by_key(db, user_id)
    return _serialize_student(student, {c.id: k for k, c in cats.items()})


@router.put("/api/students/{student_id}", response_model=StudentOut)
def update_student(
    student_id: UUID,
    body: StudentUpdate,
    user_id: Annotated[UUID, Depends(require_user_id)],
    db: Annotated[Session, Depends(get_db)],
) -> StudentOut:
    student = _get_owned_student(db, user_id, student_id)
    if body.classroom_id is not None and body.classroom_id != student.classroom_id:
        _get_owned_classroom(db, user_id, body.classroom_id)  # ensures ownership
        student.classroom_id = body.classroom_id
    student.seat_number = body.seat_number
    student.name = body.name
    student.email = body.email
    try:
        db.flush()
    except IntegrityError:
        db.rollback()
        raise _conflict(
            "errors.student.duplicate_seat",
            "Seat number already exists in this classroom.",
        )
    if body.standards is not None:
        _apply_standards(db, user_id, student, body.standards)
    db.commit()
    db.refresh(student)
    cats = _categories_by_key(db, user_id)
    return _serialize_student(student, {c.id: k for k, c in cats.items()})


@router.delete("/api/students/{student_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_student(
    student_id: UUID,
    user_id: Annotated[UUID, Depends(require_user_id)],
    db: Annotated[Session, Depends(get_db)],
) -> None:
    student = _get_owned_student(db, user_id, student_id)
    db.delete(student)
    db.commit()


# ---------- standards helper ----------

def _apply_standards(
    db: Session,
    user_id: UUID,
    student: Student,
    standards: dict[str, float],
) -> None:
    """Upsert standards passed as {system_key: threshold}.

    Unknown system_keys are silently ignored (caller validates earlier).
    Existing rows for keys not present in the payload are left alone.
    """
    if not standards:
        return
    cats = _categories_by_key(db, user_id)
    existing = {
        std.category_id: std
        for std in db.query(StudentStandard)
        .filter(StudentStandard.student_id == student.id)
        .all()
    }
    for key, value in standards.items():
        cat = cats.get(key)
        if cat is None:
            continue
        std = existing.get(cat.id)
        if std is None:
            db.add(
                StudentStandard(
                    user_id=user_id,
                    student_id=student.id,
                    category_id=cat.id,
                    threshold=Decimal(str(value)),
                )
            )
        else:
            std.threshold = Decimal(str(value))


# ---------- Excel template ----------

# Excel layout: the first 3 columns describe the student (header text in
# Row 1, data starts at Row 6). The remaining columns are score columns —
# each column declares an exam item via 4 metadata rows above the scores:
#   Row 1: 科目         (e.g., 國語 / Chinese)
#   Row 2: 類別         (e.g., 段考 / Major Exam — restricted to 3 values)
#   Row 3: 日期 選填    (e.g., 2026-05-13; falls back to today)
#   Row 4: 考試名稱 選填 (e.g., 期中考; falls back to "<category>-<date>")
#   Row 5: 分數標題    ("分數" / "Score" — visual cue only)
#   Row 6+: scores

_TPL_STUDENT_HEADERS = ("座號", "姓名（選填）", "email（選填）")
# Data rows start at row 5. Rows 1-4 of score columns hold (in order):
#   科目 / 類別 / 日期（選填）/ 考試名稱（選填）
_DATA_ROW = 5


@router.get(
    "/api/classrooms/{classroom_id}/students/template.xlsx",
    response_class=StreamingResponse,
)
def download_template(
    classroom_id: UUID,
    user_id: Annotated[UUID, Depends(require_user_id)],
    db: Annotated[Session, Depends(get_db)],
) -> StreamingResponse:
    _get_owned_classroom(db, user_id, classroom_id)
    wb = Workbook()
    ws = wb.active
    ws.title = "students"

    # Row 1: student column headers (A-C) + subject names for score columns (D+)
    ws.cell(row=1, column=1, value=_TPL_STUDENT_HEADERS[0])
    ws.cell(row=1, column=2, value=_TPL_STUDENT_HEADERS[1])
    ws.cell(row=1, column=3, value=_TPL_STUDENT_HEADERS[2])
    ws.cell(row=1, column=4, value="國語")
    ws.cell(row=1, column=5, value="數學")
    # Row 2: category (restricted to dropdown)
    ws.cell(row=2, column=4, value="段考")
    ws.cell(row=2, column=5, value="小考")
    # Row 3: date (optional)
    ws.cell(row=3, column=4, value="2026-05-13")
    # Row 4: exam name (optional; defaults to "<category>-<date>")
    ws.cell(row=4, column=4, value="期中考")
    ws.cell(row=4, column=5, value="第3週小考")

    # Example student rows (row 6+); teacher deletes before uploading
    ws.cell(row=_DATA_ROW, column=1, value=1)
    ws.cell(row=_DATA_ROW, column=2, value="範例學生甲")
    ws.cell(row=_DATA_ROW, column=4, value=80)
    ws.cell(row=_DATA_ROW, column=5, value=75)
    ws.cell(row=_DATA_ROW + 1, column=1, value=2)
    ws.cell(row=_DATA_ROW + 1, column=2, value="範例學生乙")
    ws.cell(row=_DATA_ROW + 1, column=4, value=88)

    # Data validation: category row (row 2) for score columns must be one of
    # the three allowed values. Apply broadly to D2:Z2 so teachers can add more
    # columns without re-applying validation.
    category_choices = '"段考,小考,作業"'
    dv_cat = DataValidation(
        type="list",
        formula1=category_choices,
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="無效的類別",
        error="只能選 段考 / 小考 / 作業",
    )
    dv_cat.add("D2:Z2")
    ws.add_data_validation(dv_cat)

    # Subject row (row 1) — restrict to the built-in subject names.
    subject_choices = '"國語,數學,英文,自然,社會,音樂,美術,體育,綜合"'
    dv_sub = DataValidation(
        type="list",
        formula1=subject_choices,
        allow_blank=True,
        showErrorMessage=False,
    )
    dv_sub.add("D1:Z1")
    ws.add_data_validation(dv_sub)

    # Friendly column widths
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 26
    for col_letter in ("D", "E", "F", "G", "H"):
        ws.column_dimensions[col_letter].width = 16

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": (
                'attachment; filename="students_template.xlsx"'
            ),
        },
    )


# ---------- Excel import: parse + commit ----------

@router.post(
    "/api/classrooms/{classroom_id}/students/import",
    response_model=ImportResult,
)
async def import_students(
    classroom_id: UUID,
    user_id: Annotated[UUID, Depends(require_user_id)],
    db: Annotated[Session, Depends(get_db)],
    file: Annotated[UploadFile, File(...)],
    dry_run: Annotated[bool, Query()] = True,
) -> ImportResult:
    classroom = _get_owned_classroom(db, user_id, classroom_id)

    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise _bad_request("errors.import.bad_file_type", "File must be .xlsx")

    content = await file.read()
    try:
        wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise _bad_request("errors.import.unreadable", "Excel file could not be parsed.")
    ws = wb.active
    if ws is None:
        raise _bad_request("errors.import.empty", "Workbook is empty.")

    # Validate that the student-column headers are where we expect them.
    header_errors = _check_student_headers(ws)
    if header_errors:
        raise _bad_request("errors.import.bad_header", "; ".join(header_errors))

    current_semester = (
        db.query(Semester)
        .filter(Semester.user_id == user_id, Semester.is_current.is_(True))
        .one_or_none()
    )
    if current_semester is None:
        raise _bad_request(
            "errors.import.no_current_semester",
            "No current semester is set. Configure it before importing grades.",
        )

    # Pre-load reference data
    existing_students = (
        db.query(Student)
        .filter(Student.classroom_id == classroom_id, Student.user_id == user_id)
        .all()
    )
    by_seat: dict[int, Student] = {s.seat_number: s for s in existing_students}

    subjects_by_key: dict[str, Subject] = {
        s.system_key: s
        for s in db.query(Subject)
        .filter(Subject.user_id.is_(None), Subject.system_key.isnot(None))
        .all()
    }
    categories_by_key: dict[str, Category] = {
        c.system_key: c for c in db.query(Category).filter(Category.user_id == user_id).all()
    }

    # Parse metadata for each score column (D onward; column index 3+)
    columns = _parse_score_columns(
        ws, subjects_by_key, categories_by_key, current_semester, user_id, db,
    )

    # Parse data rows starting at row 6
    student_rows = _parse_student_rows(ws, by_seat, columns)

    # Count grade-level intent
    grade_create = 0
    grade_overwrite = 0
    for srow in student_rows:
        if srow.action == "error":
            continue
        for col_idx, _score in srow.scores.items():
            col = columns[col_idx - 3] if col_idx - 3 < len(columns) else None
            if col is None or col.errors:
                continue
            existing_grade = _existing_grade(db, col.existing_item_id, srow.existing_id)
            if existing_grade is not None:
                grade_overwrite += 1
            else:
                grade_create += 1

    error_count = sum(1 for c in columns if c.errors) + sum(
        1 for r in student_rows if r.action == "error"
    )

    summary = ImportPreviewSummary(
        student_total=len(student_rows),
        student_create=sum(1 for r in student_rows if r.action == "create"),
        student_update=sum(1 for r in student_rows if r.action == "update"),
        item_total=len(columns),
        item_create=sum(1 for c in columns if not c.reuses_existing and not c.errors),
        item_reuse=sum(1 for c in columns if c.reuses_existing and not c.errors),
        grade_total=grade_create + grade_overwrite,
        grade_create=grade_create,
        grade_overwrite=grade_overwrite,
        errors=error_count,
    )

    if dry_run:
        return ImportResult(
            dry_run=True, summary=summary, columns=columns, students=student_rows,
        )

    if error_count > 0:
        raise _bad_request(
            "errors.import.has_errors",
            "Cannot import while preview contains errors.",
        )

    _commit_import(
        db, user_id, classroom, columns, student_rows, by_seat, current_semester,
    )
    db.commit()
    return ImportResult(
        dry_run=False, summary=summary, columns=columns, students=student_rows,
    )


# ---------- parse helpers ----------

def _check_student_headers(ws: Any) -> list[str]:
    """座號 must be at A1; the other two are checked loosely."""
    h1 = ws.cell(row=1, column=1).value
    if h1 is None or str(h1).strip() != "座號":
        return ["A1 must be 「座號」"]
    return []


def _parse_score_columns(
    ws: Any,
    subjects_by_key: dict[str, Subject],
    categories_by_key: dict[str, Category],
    semester: Semester,
    user_id: UUID,
    db: Session,
) -> list[ImportColumnPreview]:
    """Walk columns D..end, each column = one future Item."""
    result: list[ImportColumnPreview] = []
    max_col = ws.max_column or 0
    # Score columns start at column index 4 (column D), 0-based index 3.
    for col in range(4, max_col + 1):
        subject_raw = ws.cell(row=1, column=col).value
        category_raw = ws.cell(row=2, column=col).value
        # Skip entirely-blank score columns
        if (subject_raw in (None, "")) and (category_raw in (None, "")):
            # but only if no scores either — check briefly
            has_scores = any(
                ws.cell(row=r, column=col).value not in (None, "")
                for r in range(_DATA_ROW, (ws.max_row or _DATA_ROW) + 1)
            )
            if not has_scores:
                continue

        errors: list[str] = []
        subject_str = str(subject_raw).strip() if subject_raw not in (None, "") else None
        category_str = str(category_raw).strip() if category_raw not in (None, "") else None

        subject_key = SUBJECT_NAME_TO_KEY.get(subject_str) if subject_str else None
        category_key = CATEGORY_NAME_TO_KEY.get(category_str) if category_str else None

        if subject_str is None:
            errors.append("科目 is required for score columns")
        elif subject_key is None:
            errors.append(f"Unknown 科目: {subject_str}")
        elif subject_key not in subjects_by_key:
            errors.append(f"科目 {subject_str} not seeded — run /api/me/seed first")

        if category_str is None:
            errors.append("類別 is required for score columns")
        elif category_key is None:
            errors.append(
                f"Unknown 類別: {category_str} (allowed: 段考 / 小考 / 作業)"
            )
        elif category_key not in categories_by_key:
            errors.append(f"類別 {category_str} not seeded for user")

        # Date (row 3) — optional, falls back to today
        date_raw = ws.cell(row=3, column=col).value
        exam_date = _coerce_date(date_raw)
        if date_raw not in (None, "") and exam_date is None:
            errors.append(f"日期格式無法解析: {date_raw!r}")
        if exam_date is None:
            exam_date = date.today()

        # Name (row 4) — optional, falls back to "<category>-<date>"
        name_raw = ws.cell(row=4, column=col).value
        if name_raw not in (None, ""):
            exam_name = str(name_raw).strip()
        else:
            cat_label = category_str or ""
            exam_name = f"{cat_label}-{exam_date.isoformat()}"

        # Look up existing item (only if all keys resolve)
        existing_item_id: UUID | None = None
        reuses = False
        if (
            not errors
            and subject_key
            and category_key
            and subject_key in subjects_by_key
            and category_key in categories_by_key
        ):
            existing_item = (
                db.query(Item)
                .filter(
                    Item.user_id == user_id,
                    Item.subject_id == subjects_by_key[subject_key].id,
                    Item.category_id == categories_by_key[category_key].id,
                    Item.semester_id == semester.id,
                    Item.name == exam_name,
                )
                .one_or_none()
            )
            if existing_item is not None:
                existing_item_id = existing_item.id
                reuses = True

        result.append(
            ImportColumnPreview(
                column_index=col - 1,  # 0-based; D = 3
                subject_input=subject_str,
                subject_system_key=subject_key,
                category_input=category_str,
                category_system_key=category_key,
                exam_date=exam_date,
                exam_name=exam_name,
                existing_item_id=existing_item_id,
                reuses_existing=reuses,
                errors=errors,
            )
        )
    return result


def _parse_student_rows(
    ws: Any,
    by_seat: dict[int, Student],
    columns: list[ImportColumnPreview],
) -> list[ImportStudentRow]:
    """Walk rows 6..end. Each row = one student + optional scores."""
    rows: list[ImportStudentRow] = []
    seen_seats: set[int] = set()
    max_row = ws.max_row or 0
    valid_col_indexes = {c.column_index for c in columns if not c.errors}

    for r in range(_DATA_ROW, max_row + 1):
        # If every cell in the row is blank, skip
        row_cells = [ws.cell(row=r, column=col).value for col in range(1, (ws.max_column or 0) + 1)]
        if all(v in (None, "") for v in row_cells):
            continue

        errors: list[str] = []

        # seat
        seat_raw = ws.cell(row=r, column=1).value
        seat: int | None = None
        try:
            if seat_raw in (None, ""):
                errors.append("座號 is required")
            else:
                seat = int(seat_raw)
                if seat < 1 or seat > 99:
                    errors.append("座號 must be 1–99")
                    seat = None
        except (TypeError, ValueError):
            errors.append("座號 must be an integer")

        if seat is not None:
            if seat in seen_seats:
                errors.append("座號 duplicated in file")
            seen_seats.add(seat)

        # name + email
        name_raw = ws.cell(row=r, column=2).value
        name = str(name_raw).strip() if name_raw not in (None, "") else None
        email_raw = ws.cell(row=r, column=3).value
        email = str(email_raw).strip() if email_raw not in (None, "") else None
        if email and "@" not in email:
            errors.append("email is invalid")

        # scores per score column
        scores: dict[int, float] = {}
        for col in columns:
            if col.errors:
                continue
            cell_val = ws.cell(row=r, column=col.column_index + 1).value
            if cell_val in (None, ""):
                continue
            try:
                num = float(Decimal(str(cell_val)))
            except (InvalidOperation, ValueError):
                errors.append(f"分數欄 (col {col.column_index + 1}) is not a number")
                continue
            if num < 0 or num > 100:
                errors.append(f"分數欄 (col {col.column_index + 1}) must be 0–100")
                continue
            scores[col.column_index] = num

        existing = by_seat.get(seat) if seat is not None else None
        action: Literal["create", "update", "error"] = (
            "error" if errors else ("update" if existing else "create")
        )
        rows.append(
            ImportStudentRow(
                row_number=r,
                action=action,
                seat_number=seat,
                name=name,
                email=email,
                scores=scores,
                existing_id=existing.id if existing else None,
                errors=errors,
            )
        )
    _ = valid_col_indexes  # currently unused; kept for future cross-checks
    return rows


def _coerce_date(raw: Any) -> date | None:
    if raw is None or raw == "":
        return None
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    # YYYYMMDD integer like 20260513
    if isinstance(raw, int) and 19000101 <= raw <= 99991231:
        s = str(raw)
        try:
            return date(int(s[0:4]), int(s[4:6]), int(s[6:8]))
        except ValueError:
            return None
    if isinstance(raw, str):
        s = raw.strip().replace("/", "-")
        for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%Y%m%d"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
    return None


def _existing_grade(
    db: Session, item_id: UUID | None, student_id: UUID | None,
) -> Grade | None:
    if item_id is None or student_id is None:
        return None
    return (
        db.query(Grade)
        .filter(Grade.item_id == item_id, Grade.student_id == student_id)
        .one_or_none()
    )


# ---------- commit ----------

def _commit_import(
    db: Session,
    user_id: UUID,
    classroom: Classroom,
    columns: list[ImportColumnPreview],
    student_rows: list[ImportStudentRow],
    by_seat: dict[int, Student],
    semester: Semester,
) -> None:
    # 1. Upsert students; track seat → Student so we can link grades.
    seat_to_student: dict[int, Student] = {}
    for r in student_rows:
        if r.action == "error" or r.seat_number is None:
            continue
        if r.existing_id is not None:
            student = by_seat[r.seat_number]
            student.name = r.name
            student.email = r.email
        else:
            student = Student(
                user_id=user_id,
                classroom_id=classroom.id,
                seat_number=r.seat_number,
                name=r.name,
                email=r.email,
                source="manual",
            )
            db.add(student)
            db.flush()
        seat_to_student[r.seat_number] = student
    db.flush()

    # 2. Upsert items per column; pre-fetch system subject + user category.
    subjects_by_key: dict[str, Subject] = {
        s.system_key: s
        for s in db.query(Subject)
        .filter(Subject.user_id.is_(None), Subject.system_key.isnot(None))
        .all()
    }
    categories_by_key: dict[str, Category] = {
        c.system_key: c
        for c in db.query(Category).filter(Category.user_id == user_id).all()
    }

    col_to_item: dict[int, Item] = {}
    for col in columns:
        if col.errors:
            continue
        if col.existing_item_id is not None:
            item = db.query(Item).filter(Item.id == col.existing_item_id).one()
        else:
            subj = subjects_by_key[col.subject_system_key]  # type: ignore[index]
            cat = categories_by_key[col.category_system_key]  # type: ignore[index]
            item = Item(
                user_id=user_id,
                subject_id=subj.id,
                category_id=cat.id,
                semester_id=semester.id,
                name=col.exam_name,
            )
            db.add(item)
            db.flush()
        # Link item to this classroom if not already linked
        if classroom not in item.classrooms:
            item.classrooms.append(classroom)
        col_to_item[col.column_index] = item
    db.flush()

    # 3. Upsert grades per (student, item).
    for r in student_rows:
        if r.action == "error" or r.seat_number is None:
            continue
        student = seat_to_student.get(r.seat_number)
        if student is None:
            continue
        for col_idx, score in r.scores.items():
            item = col_to_item.get(col_idx)
            if item is None:
                continue
            existing = (
                db.query(Grade)
                .filter(Grade.item_id == item.id, Grade.student_id == student.id)
                .one_or_none()
            )
            if existing is not None:
                existing.score = Decimal(str(score))
            else:
                db.add(
                    Grade(
                        user_id=user_id,
                        item_id=item.id,
                        student_id=student.id,
                        score=Decimal(str(score)),
                        source="manual",
                    )
                )
